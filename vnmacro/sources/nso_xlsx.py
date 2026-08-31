"""Parse the NSO monthly "Biểu" workbook into tidy observations.

Every monthly report ships one .xlsx with ~20 sheets (agriculture, IIP,
enterprises, investment, FDI, retail, exports, imports, CPI, transport,
tourism...). They share a house layout:

  * a title row, then a stacked multi-row header band with merged cells;
  * a unit marker somewhere in the header ("%", "Tỷ đồng", "Triệu USD");
  * two or three left-hand columns holding a *hierarchy* of row labels
    (section in col A, item in col B, sub-item in col C).

So the parser is generic — unmerge, find where the numbers start, fold the
header band into one label per column, walk the rows — and nothing is dropped.
A mapping layer on top promotes the headline series to stable ids; everything
else keeps its Vietnamese label and gets a deterministic generated id.
"""
from __future__ import annotations

import datetime as dt
import logging
import re
from pathlib import Path

import openpyxl

from ..util import norm_ws, slugify, strip_accents, to_float

log = logging.getLogger(__name__)

MAX_HEADER_ROWS = 14
MAX_LABEL_COLS = 4

# Standalone unit markers that NSO drops into the header band, usually
# right-aligned over a single column. They belong to the sheet, not to the
# column they happen to sit above, so they are stripped out of column labels.
UNIT_TOKENS: list[tuple[str, str]] = [
    (r"nghin ty dong", "thousand billion VND"),
    (r"ty dong", "billion VND"),
    (r"trieu usd", "million USD"),
    (r"ty usd", "billion USD"),
    (r"nghin tan", "thousand tonnes"),
    (r"trieu tan", "million tonnes"),
    (r"nghin ha", "thousand ha"),
    (r"nghin m3", "thousand m3"),
    (r"trieu cay", "million trees"),
    (r"nghin luot nguoi|nghin luot", "thousand arrivals"),
    (r"trieu luot nguoi|trieu luot", "million arrivals"),
    (r"nghin nguoi", "thousand persons"),
    (r"trieu tan\.km|trieu tan km", "million tonne-km"),
    # NB: bare words like "dự án" / "doanh nghiệp" / "người" are deliberately
    # absent — they are ordinary column wording ("Số dự án"), not unit notes.
]

# Column wording that determines the unit regardless of the sheet marker.
MEASURE_UNITS: list[tuple[str, str]] = [
    (r"so voi|\(%\)|co cau|ty trong|ty le|bang \d|dat \d+%|ke hoach", "%"),
    (r"\bluong\b", "thousand tonnes"),
    (r"\btri gia\b", "million USD"),
    (r"\bso du an\b|\bso doanh nghiep\b", "count"),
]


def _detect_unit(text: str) -> str | None:
    """Pick a unit out of free header text."""
    t = strip_accents(text)
    for pat, unit in UNIT_TOKENS:
        if re.search(rf"(?<![a-z]){pat}(?![a-z])", t):
            return unit
    if "%" in (text or ""):
        return "%"
    return None


def _is_unit_marker(cell: str) -> bool:
    """True when a header cell is *only* a unit note ('Triệu USD', '%').

    These sit right-aligned over one arbitrary column, so folding them into
    that column's label would mislabel the series.
    """
    t = norm_ws(cell)
    if not t or len(t) > 34:
        return False
    if to_float(t) is not None:
        return False                      # a bare year/number is header content
    # Match on the accent-folded form: Vietnamese diacritics vary in ways a
    # hand-written character class gets wrong ("với" is ớ, not ơ).
    folded = strip_accents(t)
    if re.search(r"so voi|thang|nam|quy|cung k|binh quan|luy ke", folded):
        return False
    stripped = re.sub(r"[%;()\s\d\.\-–]", "", folded)
    if not stripped:
        return True                       # bare "%" or "(%)"
    return _detect_unit(t) is not None


def _resolve_unit(measure: str, sheet_unit: str | None) -> str | None:
    m = strip_accents(measure)
    for pat, unit in MEASURE_UNITS:
        if re.search(pat, m):
            return unit
    return sheet_unit


def _grid(ws) -> list[list]:
    """Materialise the sheet with merged cells filled across their range."""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    for rng in ws.merged_cells.ranges:
        r0, r1 = rng.min_row - 1, rng.max_row - 1
        c0, c1 = rng.min_col - 1, rng.max_col - 1
        if r0 >= len(rows):
            continue
        val = rows[r0][c0] if c0 < len(rows[r0]) else None
        if val is None:
            continue
        for r in range(r0, min(r1 + 1, len(rows))):
            for c in range(c0, min(c1 + 1, len(rows[r]))):
                if rows[r][c] is None:
                    rows[r][c] = val
    return rows


def _numeric_density(rows: list[list], col: int) -> float:
    seen = hits = 0
    for r in rows:
        if col >= len(r) or r[col] is None:
            continue
        seen += 1
        if to_float(r[col]) is not None:
            hits += 1
    return hits / seen if seen else 0.0


def _split_columns(rows: list[list]) -> tuple[list[int], list[int]]:
    """Left-hand label columns vs numeric data columns."""
    ncols = max((len(r) for r in rows), default=0)
    density = [_numeric_density(rows, c) for c in range(ncols)]
    label_cols: list[int] = []
    for c in range(min(MAX_LABEL_COLS, ncols)):
        if density[c] < 0.5:
            label_cols.append(c)
        else:
            break
    if not label_cols:
        label_cols = [0]
    data_cols = [c for c in range(ncols) if c > max(label_cols) and density[c] > 0.3]
    return label_cols, data_cols


def _first_data_row(rows: list[list], label_cols, data_cols) -> int:
    for i, r in enumerate(rows):
        if not any(norm_ws(r[c]) for c in label_cols if c < len(r)):
            continue
        if any(c < len(r) and to_float(r[c]) is not None for c in data_cols):
            return i
    return len(rows)


def _column_labels(rows, header_rows, data_cols) -> dict[int, str]:
    """Fold the stacked header band into one label per data column."""
    out = {}
    for c in data_cols:
        parts, seen = [], set()
        for r in header_rows:
            if c >= len(rows[r]):
                continue
            v = norm_ws(rows[r][c])
            if not v or v in seen or _is_unit_marker(v):
                continue
            # keep bare years ("2026") — they complete "Tháng 7 / năm / 2026"
            if to_float(v) is not None and not re.fullmatch(r"(19|20)\d{2}", v):
                continue
            parts.append(v)
            seen.add(v)
        out[c] = norm_ws(" ".join(parts))
    return out


def parse_sheet(ws, *, sheet_name: str) -> list[dict]:
    rows = _grid(ws)
    if not rows:
        return []
    label_cols, data_cols = _split_columns(rows)
    if not data_cols:
        return []
    start = _first_data_row(rows, label_cols, data_cols)
    header_rows = list(range(0, min(start, MAX_HEADER_ROWS)))
    col_labels = _column_labels(rows, header_rows, data_cols)

    # The sheet-level unit comes only from cells that are pure unit markers.
    sheet_unit = None
    for r in header_rows:
        for v in rows[r]:
            if v is not None and _is_unit_marker(norm_ws(v)):
                sheet_unit = _detect_unit(norm_ws(v)) or sheet_unit
                if sheet_unit and sheet_unit != "%":
                    break
        if sheet_unit and sheet_unit != "%":
            break

    records: list[dict] = []
    path_stack: dict[int, str] = {}
    for i in range(start, len(rows)):
        row = rows[i]
        labels = {c: norm_ws(row[c]) if c < len(row) else "" for c in label_cols}
        if any(labels.values()):
            for c in label_cols:
                if labels[c]:
                    path_stack[c] = labels[c]
                    for deeper in label_cols:
                        if deeper > c:
                            path_stack.pop(deeper, None)
        item = ""
        for c in sorted(label_cols, reverse=True):
            if labels.get(c):
                item = labels[c]
                break
        if not item:
            continue
        if strip_accents(item).startswith("trong do") and len(item) < 14:
            continue                                   # "Trong đó:" spacer
        row_path = " > ".join(path_stack[c] for c in sorted(path_stack) if path_stack[c])

        for c in data_cols:
            val = to_float(row[c]) if c < len(row) else None
            if val is None:
                continue
            measure = col_labels.get(c) or f"col{c}"
            # "Kỳ gốc 2024" names the CPI base period; it moves every rebasing
            # (2014 -> 2019 -> 2024), which is exactly what breaks naive splices.
            bm = re.search(r"ky goc\s*(\d{4})", strip_accents(measure))
            records.append({
                "sheet": sheet_name,
                "row_label": item,
                "row_path": row_path,
                "measure": measure,
                "value": val,
                "unit": _resolve_unit(measure, sheet_unit),
                "base_year": int(bm.group(1)) if bm else None,
                "col_index": c,
            })
    return records


class _LegacySheet:
    """Minimal openpyxl-compatible view over a legacy .xls worksheet.

    Reports before ~2017 ship BIFF .xls, which openpyxl cannot open at all —
    and losing them puts real holes in the CPI chain. xlrd 2.x reads exactly
    this format, so the few attributes `parse_sheet` needs are adapted here
    rather than duplicating the parser.
    """

    def __init__(self, sheet):
        self._s = sheet
        self.merged_cells = type("M", (), {"ranges": [
            type("R", (), {"min_row": r0 + 1, "max_row": r1, "min_col": c0 + 1,
                           "max_col": c1})()
            for r0, r1, c0, c1 in sheet.merged_cells
        ]})()

    def iter_rows(self, values_only: bool = True):
        for r in range(self._s.nrows):
            yield [c.value if c.value != "" else None for c in self._s.row(r)]

    def calculate_dimension(self):
        return f"1:{self._s.nrows}"


def _legacy_workbook(path: Path):
    import xlrd

    book = xlrd.open_workbook(str(path), formatting_info=False)
    return {name: _LegacySheet(book.sheet_by_name(name))
            for name in book.sheet_names()}


def parse_workbook(path: Path, *, sheets: list[str] | None = None) -> list[dict]:
    out = []
    if path.suffix.lower() == ".xls":
        try:
            book = _legacy_workbook(path)
        except Exception as exc:
            log.error("legacy .xls %s unreadable: %s", path.name, exc)
            return []
        items = book.items()
        close = None
    else:
        wb = openpyxl.load_workbook(path, data_only=True)
        items = ((name, wb[name]) for name in wb.sheetnames)
        close = wb.close

    for name, ws in items:
        if sheets and name not in sheets:
            continue
        try:
            out.extend(parse_sheet(ws, sheet_name=name))
        except Exception as exc:
            log.error("sheet %r in %s failed: %s", name, path.name, exc)
    if close:
        close()
    return out


# ------------------------------------------------------------- canonicalise --

def sheet_key(sheet_name: str) -> str:
    """'14. XK tháng' -> 'xk_thang' (stable across months despite renames)."""
    s = strip_accents(sheet_name)
    s = re.sub(r"^\s*\d+[\.\s]*", "", s)
    return slugify(s, 32).replace("-", "_")


def to_observations(raw: list[dict], *, release: dict, raw_file: str) -> list[dict]:
    """Attach release metadata and build ids.

    Dated to the release's *monthly* reference (see util._mk) because these
    sheets are monthly even in a quarterly issue; the release period itself is
    kept in ``dims`` so a row can always be traced back to its publication.
    """
    month_date = release.get("month_date") or release["period_date"]
    if isinstance(month_date, str):
        month_date = dt.date.fromisoformat(month_date)
    vintage = release.get("wp_date", "")[:10]
    vintage = dt.date.fromisoformat(vintage) if vintage else None

    obs = []
    for r in raw:
        sk = sheet_key(r["sheet"])
        sid = f"NSO.{sk}.{slugify(r['row_label'], 40)}.{slugify(r['measure'], 40)}"
        dims = {"sheet": r["sheet"], "row_path": r["row_path"], "col": r["col_index"],
                "release_period": release.get("ref_period")}
        if r.get("base_year"):
            dims["cpi_base_year"] = r["base_year"]
        obs.append({
            "series_id": sid,
            "dataset": "nso_monthly_tables",
            "source": "NSO",
            "freq": "M",
            "date": month_date,
            "ref_period": release.get("month_ref") or release["ref_period"],
            "value": r["value"],
            "unit": r["unit"],
            "scale": 0,
            "status": "so_bo",
            "vintage": vintage,
            "partner": None,
            "breakdown": r["row_path"] or r["row_label"],
            "label_vi": r["row_label"],
            "measure": r["measure"],
            "dims": dims,
            "raw_file": raw_file,
        })
    return obs
